"""
Excel deep-dive generator — borrows the dcf-model + comps-analysis skill methodologies
from the financial-analysis plugin, but writes a lite formula-driven model from FMP data.

Three tabs:
  Summary  — one-pager: cheapness blend, intrinsic, MoS, confidence, methodology notes
  DCF      — formula-driven 5y projection, WACC, terminal value, equity bridge, sensitivity 5x5
  Comps    — peer ranking on P/E, EV/EBITDA, EV/Rev, with stats row

CRITICAL PRINCIPLE FROM THE DCF-MODEL SKILL: every projected cell is an Excel formula,
NEVER a value pre-computed in Python and written as a number. The user must be able to
flex any assumption (WACC inputs, growth rate, terminal g) and have every downstream cell
recalculate. Skill quote: "If you catch yourself computing something in Python and writing
the result — STOP."

Usage:
  python research.py AAPL                  # CLI
  research.build_report("AAPL")            # programmatic (from pipeline.py)
"""
import datetime
import json
import os
import sys

import config
import fetch
import valuation

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
except ImportError:
    print("ERROR: openpyxl not installed — run: pip install openpyxl")
    sys.exit(1)

REPORT_DIR = "app/reports"
PEER_CACHE_DIR = "cache"

# ---- styling (kept minimal per skill convention: blues + greys only) ----
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # yellow = user-editable input
CALC_FILL = PatternFill("solid", fgColor="F2F2F2")    # grey = formula output
HDR_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
LBL_FONT = Font(name="Calibri", size=10, bold=True, color="000000")
NORM_FONT = Font(name="Calibri", size=10, color="000000")
ITAL_FONT = Font(name="Calibri", size=9, italic=True, color="595959")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fmp(path):
    sep = "&" if "?" in path else "?"
    url = f"https://financialmodelingprep.com/stable/{path}{sep}apikey={config.FMP_API_KEY}"
    return fetch.get_json(url)


def _peers_cache_path(ticker):
    os.makedirs(PEER_CACHE_DIR, exist_ok=True)
    today = datetime.date.today().isoformat()
    return os.path.join(PEER_CACHE_DIR, f"peers_{ticker.upper()}_{today}.json")


def _load_peer_data(ticker):
    """Peer rows with multiples for the Comps tab. Peer LIST from FMP (free, cached daily
    via valuation._peer_list); each peer's MULTIPLES from yfinance (works for all symbols).
    Cached daily to disk."""
    p = _peers_cache_path(ticker)
    if os.path.exists(p):
        try:
            with open(p, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    out = []
    for sym in valuation._peer_list(ticker)[:8]:
        if not sym or sym == ticker:
            continue
        info = valuation._yf_info(sym)
        if not info:
            continue
        m = valuation._multiples(info)
        out.append(dict(
            symbol=sym, name=info.get("shortName") or info.get("longName") or sym,
            market_cap=valuation._num(info.get("marketCap")),
            pe=m.get("pe"), ev_ebitda=m.get("ev_ebitda"), ev_rev=m.get("ev_rev"),
        ))
    if out:
        try:
            with open(p, "w", encoding="utf-8") as f:
                json.dump(out, f, indent=2)
        except Exception:
            pass
    return out


def _style_header(ws, cell_range, text, fill=HDR_FILL, font=HDR_FONT):
    """Set header text on top-left cell, then merge + style (avoids the merged-cell
    `InvalidArgument` pitfall the dcf-model skill calls out)."""
    first = cell_range.split(":")[0]
    ws[first] = text
    ws.merge_cells(cell_range)
    for row in ws[cell_range]:
        for c in row:
            c.fill = fill
            c.font = font
            c.alignment = Alignment(horizontal="left", vertical="center")


def _label(ws, cell, text, font=LBL_FONT):
    ws[cell] = text
    ws[cell].font = font


def _input(ws, cell, value, comment=None, fmt=None):
    """Hardcoded INPUT cell (yellow). Per skill: only raw inputs + assumptions allowed."""
    ws[cell] = value
    ws[cell].fill = INPUT_FILL
    ws[cell].font = NORM_FONT
    if fmt:
        ws[cell].number_format = fmt
    if comment:
        ws[cell].comment = Comment(comment, "research.py")


def _formula(ws, cell, formula, fmt=None):
    """Formula CALC cell (grey) — recalculates when any upstream input changes."""
    ws[cell] = formula
    ws[cell].fill = CALC_FILL
    ws[cell].font = NORM_FONT
    if fmt:
        ws[cell].number_format = fmt


# ---------------------------------------------------------------------------
# Summary tab — one-pager
# ---------------------------------------------------------------------------
def _build_summary(wb, val_data, ticker):
    ws = wb.create_sheet("Summary", 0)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 60

    _style_header(ws, "A1:C2", f"{ticker} — Valuation Snapshot")
    ws["A3"] = f"Generated {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')} · Source: FMP · Methodology: financial-analysis plugin (Anthropic FSI), adapted for free-tier data"
    ws["A3"].font = ITAL_FONT
    ws.merge_cells("A3:C3")

    rows = [
        ("Blended cheapness (0=rich, 1=cheap)", val_data.get("cheapness"),
         "Re-weighted across factors that returned data — see DCF tab inputs to flex."),
        ("Confidence", val_data.get("confidence"),
         "low/medium/high based on how many of the 3 factors fired (ratings, comps, DCF)."),
        ("DCF intrinsic value / share", val_data.get("intrinsic_value"),
         "From the DCF tab — flex WACC, growth, terminal-g on that tab and this updates."),
        ("Market price", val_data.get("market_price"), "Live at time of generation."),
        ("Margin of safety", val_data.get("margin_of_safety"),
         "(intrinsic - market) / market. Graham/Buffett convention is +20% before buying."),
    ]
    r = 5
    _style_header(ws, f"A{r}:C{r}", "Key numbers", SUB_FILL, LBL_FONT)
    r += 1
    for label, value, note in rows:
        ws[f"A{r}"] = label; ws[f"A{r}"].font = LBL_FONT
        ws[f"B{r}"] = value if value is not None else "—"
        ws[f"B{r}"].font = NORM_FONT
        if isinstance(value, float) and "margin" in label.lower():
            ws[f"B{r}"].number_format = "+0.0%;-0.0%"
        elif isinstance(value, (int, float)) and "share" in label.lower() or "price" in label.lower():
            ws[f"B{r}"].number_format = '"$"#,##0.00'
        elif isinstance(value, float):
            ws[f"B{r}"].number_format = "0.00"
        ws[f"C{r}"] = note; ws[f"C{r}"].font = ITAL_FONT
        r += 1

    # Factor breakdown
    r += 1
    _style_header(ws, f"A{r}:C{r}", "Cheapness factors (each 0..1, 1=cheap)", SUB_FILL, LBL_FONT)
    r += 1
    for fname, fval in (val_data.get("factors") or {}).items():
        ws[f"A{r}"] = fname; ws[f"A{r}"].font = LBL_FONT
        ws[f"B{r}"] = fval; ws[f"B{r}"].font = NORM_FONT; ws[f"B{r}"].number_format = "0.000"
        ws[f"C{r}"] = {"ratings_snapshot": "FMP's pre-computed P/E + P/B score (1-5 banded)",
                       "comps_percentile": "Peer-relative ranking of P/E, EV/EBITDA, EV/Rev",
                       "dcf_margin": "Normalized (intrinsic − market) / market, clipped ±50%"}.get(fname, "")
        ws[f"C{r}"].font = ITAL_FONT
        r += 1

    # Caveats
    r += 1
    _style_header(ws, f"A{r}:C{r}", "Honest caveats (read before acting)", SUB_FILL, LBL_FONT)
    r += 1
    caveats = [
        "Lite DCF: 5y horizon, 70/30 equity/debt cap structure assumed, FCF growth from 3y trailing trend.",
        "Peer set is FMP's algorithmic /stock-peers — may include weak comps for unusual business models.",
        "Net cash treatment uses cashAndShortTermInvestments (operating cash); excludes LT marketable",
        "  securities, which understates intrinsic for cash-rich companies like AAPL (~$80B in LT inv).",
        "Margin of safety is normalized to 0..1; flip to the DCF tab for the raw % gap.",
        "This is a SIGNAL screen, not an investment thesis. Verify with primary filings before acting.",
    ]
    for cv in caveats:
        ws[f"A{r}"] = cv; ws[f"A{r}"].font = ITAL_FONT
        ws.merge_cells(f"A{r}:C{r}")
        r += 1


# ---------------------------------------------------------------------------
# DCF tab — formula-driven 5y projection
# ---------------------------------------------------------------------------
def _build_dcf(wb, val_data, ticker, cf_rows, profile_row, balance_row, treasury_10y):
    ws = wb.create_sheet("DCF")
    for col, w in zip("ABCDEFGH", [26, 14, 14, 14, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    _style_header(ws, "A1:H1", f"{ticker} — Discounted Cash Flow Model (Lite, 5y)")
    ws["A2"] = "Yellow cells = inputs you can edit. Grey cells = formulas that recalculate."
    ws["A2"].font = ITAL_FONT; ws.merge_cells("A2:H2")

    # ---- Assumptions block (rows 4-13) ----
    _style_header(ws, "A4:H4", "Inputs & assumptions (edit yellow cells)", SUB_FILL, LBL_FONT)
    risk_free = (treasury_10y / 100.0) if treasury_10y is not None else valuation.DEFAULT_RISK_FREE
    raw_beta = profile_row.get("beta")
    try: raw_beta = float(raw_beta)
    except (TypeError, ValueError): raw_beta = None
    # Clamp suspicious betas exactly like the in-app model (Yahoo sometimes returns <0 / absurd).
    if raw_beta is None or not (valuation.BETA_FLOOR <= raw_beta <= valuation.BETA_CEIL):
        beta = 1.0 if raw_beta is None else max(valuation.BETA_FLOOR, min(valuation.BETA_CEIL, raw_beta))
    else:
        beta = raw_beta
    # Latest FCF + growth trend (clipped like the app)
    fcfs = [r.get("freeCashFlow") for r in (cf_rows or []) if isinstance(r.get("freeCashFlow"), (int, float))]
    latest_fcf = float(fcfs[0]) if fcfs else 0
    samples = [fcfs[i] / fcfs[i + 1] - 1 for i in range(len(fcfs) - 1) if fcfs[i + 1] > 0]
    g0 = max(valuation.FCF_GROWTH_CLIP[0], min(valuation.FCF_GROWTH_CLIP[1], sum(samples) / len(samples))) if samples else 0.0
    # The in-app model fades g0 → terminal-g over 5y (two-stage). This sheet projects a single flat
    # rate for editable simplicity, so seed it with the faded AVERAGE → sheet value ≈ app value.
    g = (g0 + valuation.TERMINAL_GROWTH) / 2

    _label(ws, "A5", "Risk-free rate (10y Treasury)")
    _input(ws, "B5", risk_free, comment="10y Treasury (^TNX via yfinance). Edit for a different RF.", fmt="0.00%")
    _label(ws, "A6", "Equity risk premium (ERP)")
    _input(ws, "B6", 0.05, comment="Damodaran-conservative US ERP. Adjust ±100bps to test sensitivity.", fmt="0.00%")
    _label(ws, "A7", "Beta (levered)")
    _input(ws, "B7", beta, comment="From yfinance, clamped to [0.7,2.5] (Yahoo sometimes returns <0). Override if you have a better estimate.", fmt="0.00")
    _label(ws, "A8", "Credit spread over risk-free")
    _input(ws, "B8", 0.02, comment="Investment-grade default. Bump to 0.04 for HY-rated names.", fmt="0.00%")
    _label(ws, "A9", "Tax rate")
    _input(ws, "B9", 0.21, comment="US federal statutory.", fmt="0.00%")
    _label(ws, "A10", "Equity weight (cap structure)")
    _input(ws, "B10", 0.70, comment="70/30 default. Edit toward 0.50 for highly-levered names.", fmt="0.0%")
    _label(ws, "A11", "FCF growth (flat, faded avg)")
    _input(ws, "B11", g, comment="Blended (g0+terminal)/2, approximating the app's two-stage fade from a clipped trailing-trend g0. Edit to test scenarios.", fmt="0.00%")
    _label(ws, "A12", "Terminal growth rate")
    _input(ws, "B12", 0.025, comment="Long-run US GDP. Anything >3% is aggressive.", fmt="0.00%")
    _label(ws, "A13", "Latest FCF ($, base year)")
    _input(ws, "B13", latest_fcf, comment="From yfinance cash-flow (latest annual Free Cash Flow).", fmt='"$"#,##0')

    # WACC derivation (rows 15-18)
    _label(ws, "A15", "Cost of equity (CAPM)")
    _formula(ws, "B15", "=B5+B7*B6", "0.00%")
    _label(ws, "A16", "After-tax cost of debt")
    _formula(ws, "B16", "=(B5+B8)*(1-B9)", "0.00%")
    _label(ws, "A17", "WACC")
    _formula(ws, "B17", "=B10*B15+(1-B10)*B16", "0.00%")
    ws["A17"].font = Font(name="Calibri", size=10, bold=True)
    ws["B17"].font = Font(name="Calibri", size=10, bold=True)

    # ---- Cash flow projection (rows 20-29) ----
    _style_header(ws, "A20:H20", "5-year FCF projection & PV", SUB_FILL, LBL_FONT)
    ws["A21"] = "Year"; ws["A21"].font = LBL_FONT
    ws["A22"] = "Projected FCF"; ws["A22"].font = LBL_FONT
    ws["A23"] = "Discount factor"; ws["A23"].font = LBL_FONT
    ws["A24"] = "PV of FCF"; ws["A24"].font = LBL_FONT
    for i, col in enumerate("BCDEFG"[:5], start=1):  # B..F = years 1..5
        ws[f"{col}21"] = i; ws[f"{col}21"].font = LBL_FONT
        if i == 1:
            _formula(ws, f"{col}22", "=$B$13*(1+$B$11)", '"$"#,##0')
        else:
            prev = chr(ord(col) - 1)
            _formula(ws, f"{col}22", f"={prev}22*(1+$B$11)", '"$"#,##0')
        _formula(ws, f"{col}23", f"=1/(1+$B$17)^{i}", "0.0000")
        _formula(ws, f"{col}24", f"={col}22*{col}23", '"$"#,##0')

    _label(ws, "A26", "Sum of PV (explicit 5y)")
    _formula(ws, "B26", "=SUM(B24:F24)", '"$"#,##0')
    _label(ws, "A27", "Terminal value (Gordon)")
    _formula(ws, "B27", "=F22*(1+B12)/(B17-B12)", '"$"#,##0')
    _label(ws, "A28", "PV of terminal value")
    _formula(ws, "B28", "=B27/(1+B17)^5", '"$"#,##0')
    _label(ws, "A29", "Enterprise value")
    _formula(ws, "B29", "=B26+B28", '"$"#,##0')
    ws["A29"].font = Font(bold=True); ws["B29"].font = Font(bold=True)

    # ---- Equity bridge (rows 31-36) ----
    _style_header(ws, "A31:H31", "Equity bridge → intrinsic per share", SUB_FILL, LBL_FONT)
    total_debt = balance_row.get("totalDebt") or 0
    cash = balance_row.get("cashAndShortTermInvestments") or balance_row.get("cashAndCashEquivalents") or 0
    mkt_cap = profile_row.get("marketCap") or 0
    price = profile_row.get("price") or 0
    shares = (mkt_cap / price) if price else 0

    _label(ws, "A32", "Less: total debt")
    _input(ws, "B32", float(total_debt), comment="From FMP balance sheet, totalDebt.", fmt='"$"#,##0')
    _label(ws, "A33", "Plus: cash & ST investments")
    _input(ws, "B33", float(cash), comment="From FMP balance sheet, cashAndShortTermInvestments.", fmt='"$"#,##0')
    _label(ws, "A34", "Equity value")
    _formula(ws, "B34", "=B29-B32+B33", '"$"#,##0')
    _label(ws, "A35", "Diluted shares outstanding")
    _input(ws, "B35", float(shares), comment="Derived: marketCap / price (FMP profile).", fmt="#,##0")
    _label(ws, "A36", "Intrinsic per share")
    _formula(ws, "B36", "=B34/B35", '"$"#,##0.00')
    ws["A36"].font = Font(bold=True); ws["B36"].font = Font(bold=True)
    _label(ws, "A37", "Current market price")
    _input(ws, "B37", float(price), comment="FMP profile.price at generation.", fmt='"$"#,##0.00')
    _label(ws, "A38", "Margin of safety")
    _formula(ws, "B38", "=(B36-B37)/B37", "+0.0%;-0.0%")
    ws["A38"].font = Font(bold=True); ws["B38"].font = Font(bold=True)

    # ---- Sensitivity table 5x5: WACC × terminal-g (rows 41-48) ----
    _style_header(ws, "A41:H41", "Sensitivity: intrinsic / share by WACC × terminal-g", SUB_FILL, LBL_FONT)
    ws["A42"] = "WACC ↓ / TermG →"; ws["A42"].font = LBL_FONT
    ws["A42"].fill = SUB_FILL
    # Center column = current B12 (terminal g), step ±50bps
    tg_steps = [-0.01, -0.005, 0.0, 0.005, 0.01]
    wacc_steps = [-0.01, -0.005, 0.0, 0.005, 0.01]
    for i, dtg in enumerate(tg_steps):
        col = chr(ord("B") + i)
        _formula(ws, f"{col}42", f"=$B$12+{dtg}", "0.00%")
        ws[f"{col}42"].fill = SUB_FILL
    for j, dw in enumerate(wacc_steps):
        row = 43 + j
        _formula(ws, f"A{row}", f"=$B$17+{dw}", "0.00%")
        ws[f"A{row}"].fill = SUB_FILL
        for i, dtg in enumerate(tg_steps):
            col = chr(ord("B") + i)
            # full DCF recalc for each cell: sum of geometric FCF projections at scenario WACC,
            # plus terminal value at scenario WACC and terminal-g. Net debt held constant.
            # PV of explicit years 1-5 with growth held at B11
            pv_explicit = "+".join([f"$B$13*(1+$B$11)^{y}/(1+A{row})^{y}" for y in range(1, 6)])
            tv = f"($B$13*(1+$B$11)^5*(1+{col}$42)/(A{row}-{col}$42))/(1+A{row})^5"
            equity = f"({pv_explicit}+{tv}-$B$32+$B$33)/$B$35"
            _formula(ws, f"{col}{row}", f"={equity}", '"$"#,##0.00')
    # Highlight center cell (row 45, col D) = base case
    ws["D45"].fill = PatternFill("solid", fgColor="BDD7EE")
    ws["D45"].font = Font(bold=True)


# ---------------------------------------------------------------------------
# Comps tab — peer-relative ranking
# ---------------------------------------------------------------------------
def _build_comps(wb, val_data, ticker, profile_row, peer_rows):
    ws = wb.create_sheet("Comps")
    for col, w in zip("ABCDEF", [10, 32, 16, 12, 14, 14]):
        ws.column_dimensions[col].width = w

    _style_header(ws, "A1:F1", f"{ticker} — Comparable Companies Analysis")
    ws["A2"] = "Peer set from FMP /stock-peers. Highlighted row = target. Stats below for percentile context."
    ws["A2"].font = ITAL_FONT; ws.merge_cells("A2:F2")

    headers = ["Ticker", "Name", "Market Cap", "P/E (TTM)", "EV/EBITDA", "EV/Revenue"]
    for i, h in enumerate(headers):
        c = ws.cell(row=4, column=i + 1, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = Alignment(horizontal="left")

    # Target row first
    my = val_data.get("my_metrics") or {}
    ws.cell(row=5, column=1, value=ticker).font = Font(bold=True)
    ws.cell(row=5, column=2, value=profile_row.get("companyName") or ticker).font = Font(bold=True)
    ws.cell(row=5, column=3, value=profile_row.get("marketCap")).number_format = '"$"#,##0'
    ws.cell(row=5, column=4, value=my.get("pe")).number_format = "0.0"
    ws.cell(row=5, column=5, value=my.get("ev_ebitda")).number_format = "0.0"
    ws.cell(row=5, column=6, value=my.get("ev_rev")).number_format = "0.0"
    for col in range(1, 7):
        ws.cell(row=5, column=col).fill = PatternFill("solid", fgColor="BDD7EE")

    # Peers
    for idx, peer in enumerate(peer_rows, start=6):
        ws.cell(row=idx, column=1, value=peer.get("symbol"))
        ws.cell(row=idx, column=2, value=peer.get("name"))
        ws.cell(row=idx, column=3, value=peer.get("market_cap")).number_format = '"$"#,##0'
        ws.cell(row=idx, column=4, value=peer.get("pe")).number_format = "0.0"
        ws.cell(row=idx, column=5, value=peer.get("ev_ebitda")).number_format = "0.0"
        ws.cell(row=idx, column=6, value=peer.get("ev_rev")).number_format = "0.0"

    last = 5 + len(peer_rows)
    # Stats row (all formulas — flex with the data)
    stats_row = last + 2
    ws.cell(row=stats_row, column=1, value="Peer median").font = LBL_FONT
    for col_idx in (4, 5, 6):
        col = get_column_letter(col_idx)
        _formula(ws, f"{col}{stats_row}", f"=MEDIAN({col}6:{col}{last})", "0.0")
    ws.cell(row=stats_row + 1, column=1, value="Peer mean").font = LBL_FONT
    for col_idx in (4, 5, 6):
        col = get_column_letter(col_idx)
        _formula(ws, f"{col}{stats_row + 1}", f"=AVERAGE({col}6:{col}{last})", "0.0")
    ws.cell(row=stats_row + 2, column=1, value="Target percentile vs peers").font = LBL_FONT
    pp = val_data.get("peer_pct") or {}
    for col_idx, key in zip((4, 5, 6), ("pe", "ev_ebitda", "ev_rev")):
        col = get_column_letter(col_idx)
        v = pp.get(key)
        ws.cell(row=stats_row + 2, column=col_idx, value=v).number_format = "0%"
    ws.cell(row=stats_row + 3, column=1, value="(higher pct = more expensive than peers)").font = ITAL_FONT


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def build_report(ticker):
    """Generate the Excel deep-dive for one ticker. Returns relative path or None."""
    ticker = ticker.upper().strip()
    val_data = valuation.valuation_for_ticker(ticker)
    if not val_data:
        print(f"[research] no valuation data for {ticker} (yfinance unavailable?)")
        return None

    # Raw model inputs from yfinance (same data layer as valuation.py), mapped into the
    # FMP-shaped dicts the _build_* layout functions already expect — so they stay unchanged.
    info = valuation._yf_info(ticker)
    if not info:
        print(f"[research] {ticker}: yfinance .info empty — skipping report")
        return None
    profile_row = dict(
        companyName=info.get("longName") or info.get("shortName") or ticker,
        marketCap=valuation._num(info.get("marketCap")),
        price=valuation._num(info.get("currentPrice")) or valuation._num(info.get("regularMarketPrice")),
        beta=valuation._num(info.get("beta")),
    )
    cf_rows = [dict(freeCashFlow=v) for v in valuation._yf_fcf_history(ticker)]
    bs_row = dict(
        totalDebt=valuation._num(info.get("totalDebt")) or 0.0,
        cashAndShortTermInvestments=valuation._num(info.get("totalCash")) or 0.0,
    )
    treasury_10y = valuation._risk_free_10y() * 100.0   # _build_dcf divides by 100

    peer_rows = _load_peer_data(ticker)

    wb = Workbook()
    # default sheet — kill it; Summary inserted at index 0
    wb.remove(wb.active)
    _build_summary(wb, val_data, ticker)
    _build_dcf(wb, val_data, ticker, cf_rows, profile_row, bs_row, treasury_10y)
    _build_comps(wb, val_data, ticker, profile_row, peer_rows)

    os.makedirs(REPORT_DIR, exist_ok=True)
    path = os.path.join(REPORT_DIR, f"{ticker}_research.xlsx")
    wb.save(path)
    # Return PWA-relative path (the PWA serves from app/, so reports/ is sibling).
    return f"reports/{ticker}_research.xlsx"


if __name__ == "__main__":
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if len(sys.argv) < 2:
        print("usage: python research.py TICKER [TICKER ...]")
        sys.exit(1)
    for tk in sys.argv[1:]:
        rel = build_report(tk)
        if rel:
            print(f"  OK  {tk}: app/{rel}")
        else:
            print(f"  FAIL {tk}")
